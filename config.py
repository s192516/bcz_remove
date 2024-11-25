
import json
import os
import logging
import openpyxl


class Config:

    def __init__(self):

        self.init_config()

        # 序号, share_key, 小班名字英文, token, 小班名字, 最低打卡天数, 打卡率, 最晚打卡时间
        # self.idx2info = {"1": ["1", "59z67un546wtzdn5","shujiang", "oVrf9AnYyFI4q%2B40bV0IrA%3D%3D", "蜀将何在", 0, 0.9, -1] ,
        #                  "2": ["2", "1xkqvpbwbagym88o", "zaoqi", "wLWZj49O0Xf3kp%2BGBGcgg8w7ZHyJfzxZhI6yJdCkApU%3D", "枣起",0,0.9, -1],
        #                  "3": ["3", "57dourk6jopdhzu0", "little_duck", "wLWZj49O0Xf3kp%2BGBGcgg8w7ZHyJfzxZhI6yJdCkApU%3D", "会客厅",0,0.9, -1],
        #                  "4": ["4", "66t61rxd7j6v842h", "chong", "H45Nfn2sIlLJ/BI8RLTePhb3l031Nfs3mMrc4cIUJ9I%3D", "冲冲冲",0,0.9, -1],
        #                  "5": ["5", "66itpxjmgw2cyufs", "bisheng", "wLWZj49O0Xf3kp%2BGBGcgg8w7ZHyJfzxZhI6yJdCkApU%3D", "必胜",0,0.9, -1],
        #                  "6": ["6", "48k4c8uqrlepkoea", "dujiang", "wLWZj49O0Xf3kp%2BGBGcgg8w7ZHyJfzxZhI6yJdCkApU%3D", "渡江",0,0.9, -1],
        #                  "7": ["7", "4b9rou2wqbs19rau", "moon", "NSq52H/vWdYdXl0loIk4OkpH2qe6imAxjbwXVW3J9PE%3D", "月亮派对",0,0.9, -1],
        #                  "8": ["8", "cm92xh786a00fsq6", "sun", "xy%2BwHPgTys6nUoBnBRvz%2BqhRkD6wu32WHqUvo%2Bc01Xg%3D", "扶光",0,0.9, -1],
        #                  "9": ["9", "6cw5tp61slkihon2", "flower", "NNSJuVVvglTIBgbv26PCrzPBZUEbDbiV8MBeGHJOVQ8%3D", "一束花",0,0.9, -1],
        #                  }


        self.default_dict = {
            "watch_list": "1+1",
            'database_path': 'data.db',
            'main_token': '',
            'blackID_file': './王者联盟黑名单正式版.xlsx',
            'auto_purge_list': "自动筛选人员名单.xlsx",
            'host': '127.0.0.1',
            'port': 8840,
            'database_path': 'data.db',
            'output_file': './小班数据.xlsx',
            'blackID_file': './王者联盟黑名单正式版.xlsx',
            'daily_record': '59 23 * * *',
            'daily_verify': '00 04 * * *',
            "scheduled_task_interval": 7,
            "group_rules": {
                "1": {
                    "idx": "1",
                    "share_key": "59z67un546wtzdn5",
                    "group_name": "shujiang",
                    "token": "oVrf9AnYyFI4q%2B40bV0IrA%3D%3D",
                    "name": "蜀将何在",
                    "completedTimes": 40,
                    "rate": 0.9,
                    "daka_time": 18,
                    "pupil": "1+2+3+4"
                },
                "2": {
                    "idx": "2",
                    "share_key": "1xkqvpbwbagym88o",
                    "group_name": "zaoqi",
                    "token": "wLWZj49O0Xf3kp%2BGBGcgg8w7ZHyJfzxZhI6yJdCkApU%3D",
                    "name": "枣起",
                    "completedTimes": 100,
                    "rate": 0.9,
                    "daka_time": 9,
                    "pupil": "1+2+3+4+5"

                },

                "3": {
                    "idx": "3",
                    "share_key": "57dourk6jopdhzu0",
                    "group_name": "little_duck",
                    "token": "wLWZj49O0Xf3kp%2BGBGcgg8w7ZHyJfzxZhI6yJdCkApU%3D",
                    "name": "会客厅",
                    "completedTimes": 100,
                    "rate": 0.9,
                    "daka_time": -1,
                    "pupil": "1+2+3+4"

                },
                "4": {
                    "idx": "4" ,
                    "share_key": "66t61rxd7j6v842h",
                    "group_name": "chong",
                    "token": "H45Nfn2sIlLJ/BI8RLTePhb3l031Nfs3mMrc4cIUJ9I%3D",
                    "name": "冲冲冲",
                    "completedTimes": 100,
                    "rate": 0.9,
                    "daka_time": -1,
                    "pupil": ""
                },

                "5": {
                    "idx": "5",
                    "share_key": "66itpxjmgw2cyufs",
                    "group_name": "bisheng",
                    "token": "wLWZj49O0Xf3kp%2BGBGcgg8w7ZHyJfzxZhI6yJdCkApU%3D",
                    "name": "必胜",
                    "completedTimes": 100,
                    "rate": 0.9,
                    "daka_time": -1,
                    "pupil": ""

                },
                "6": {
                    "idx": "6",
                    "share_key": "48k4c8uqrlepkoea",
                    "group_name": "dujiang",
                    "token": "wLWZj49O0Xf3kp%2BGBGcgg8w7ZHyJfzxZhI6yJdCkApU%3D",
                    "name": "渡江",
                    "completedTimes": 100,
                    "rate": 0.9,
                    "daka_time": -1,
                    "pupil": ""

                },
                "7": {
                    "idx": "7",
                    "share_key": "4b9rou2wqbs19rau",
                    "group_name": "moon",
                    "token": "NSq52H/vWdYdXl0loIk4OkpH2qe6imAxjbwXVW3J9PE%3D",
                    "name": "月亮派对",
                    "completedTimes": 100,
                    "rate": 0.9,
                    "daka_time": -1,
                    "pupil": "",

                },
                "8": {
                    "idx": "8",
                    "share_key": "cm92xh786a00fsq6",
                    "group_name": "sun",
                    "token": "xy%2BwHPgTys6nUoBnBRvz%2BqhRkD6wu32WHqUvo%2Bc01Xg%3D",
                    "name": "扶光",
                    "completedTimes": 100,
                    "rate": 0.9,
                    "daka_time": -1,
                    "pupil": ""

                },
                "9": {
                    "idx": "9",
                    "share_key": "6cw5tp61slkihon2",
                    "group_name": "flower",
                    "token": "NNSJuVVvglTIBgbv26PCrzPBZUEbDbiV8MBeGHJOVQ8%3D",
                    "name": "一束花",
                    "completedTimes": 100,
                    "rate": 0.9,
                    "daka_time": -1,
                    "pupil": ""

                },
                "10":{
                    "idx": "10",
                    "share_key": "f52vhq9ls5kau093",
                    "group_name": "smart",
                    "token": "%2BFA3bzPGbW3yp3yulqWXNDlU2R8afsqbNvGm9mbQgDY%3D",
                    "name": "智班",
                    "completedTimes": 50,
                    "rate": 1,
                    "daka_time": -1,

                }

            },



        }
        self.json_path = "./config.json"
        self.database_path = "./数据库.db"
        self.update_default_dict()
        self.watch_list = self.default_dict.get("watch_list")
        self.main_token = "oVrf9AnYyFI4q%2B40bV0IrA%3D%3D"
        self.auto_purge_list = self.default_dict['auto_purge_list']
        self.blacklist_path = self.default_dict['blackID_file']
        self.remove_member_interval_seconds = self.default_dict.get("scheduled_task_interval")
        self.prase_pulil()

    def init_config(self):


        # 创建一个日志记录器
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.DEBUG)

        # 创建一个用于 debug 日志的处理程序
        debug_handler = logging.FileHandler('debug.log', mode='w')
        debug_handler.setLevel(logging.DEBUG)

        # 创建一个用于 info 日志的处理程序
        info_handler = logging.FileHandler('info.log', mode='w')
        info_handler.setLevel(logging.INFO)

        # 创建一个用于 error 日志的处理程序
        error_handler = logging.FileHandler('error.log', mode='w')
        error_handler.setLevel(logging.WARNING)

        # 创建一个格式化器
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

        # 将格式化器添加到所有处理程序
        debug_handler.setFormatter(formatter)
        info_handler.setFormatter(formatter)
        error_handler.setFormatter(formatter)

        # 将处理程序添加到日志记录器
        logger.addHandler(debug_handler)
        logger.addHandler(info_handler)
        logger.addHandler(error_handler)

        self.logger = logger


    def update_default_dict(self):


        try:
            with open(self.json_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            self.default_dict["watch_list"] = data.get("watch_list")
            self.default_dict["remove_member_interval_seconds"] = data.get("scheduled_task_interval")

            for k, v in data.items():
                if k == "group_rules":
                    continue
                self.default_dict[k] = v

            for idx,rules in data.get("group_rules").items():
                for k, v in rules.items():
                    self.default_dict["group_rules"][idx][k] = v

        except:

            if not os.path.isfile(self.json_path):
                self.logger.error(  f"请检查当前文件夹不存在config.json文件")
                print(  f"请检查当前文件夹不存在config.json文件")


    def get_blacklist(self) -> set:
        blacklist = set()
        try:
            workbook = openpyxl.load_workbook(self.blacklist_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in range(2, sheet.max_row + 1):
                    bczId = sheet.cell(row=row, column=2).value
                    try:
                        blacklist.add( int(bczId))
                    except:
                        pass
        except:
            self.logger.error(f"读取{self.blacklist_path}文件失败")
        finally:
            return blacklist

    def get_watch_groups_info(self):
        groups_info = []

        watch_idxs = self.watch_list.replace(" ", "")
        watch_idxs = watch_idxs.strip()
        watch_idxs = watch_idxs.split('+')


        for idx in watch_idxs:
            group_info = self.default_dict.get("group_rules").get(idx)
            groups_info.append(group_info)

        return groups_info

    def prase_pulil(self):

        group_rules = self.default_dict['group_rules']

        for k, rules in group_rules.items():
            pupil = rules.get("pupil")
            pupil = pupil.strip()
            pupil = pupil.replace(" ", "")
            pupil = pupil.split("+")

            res = set()

            for p in pupil:
                if p == "1": res.add("一年级")
                if p == "2": res.add("二年级")
                if p == "3": res.add("三年级")
                if p == "4": res.add("四年级")
                if p == "5": res.add("五年级")
                if p == "6": res.add("六年级")
            res.add("小托福")
            res.add("KET")
            res.add("新概念第一册")
            res.add("新概念第二册")
            rules["pupil"] = res

